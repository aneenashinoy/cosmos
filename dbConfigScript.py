import openpyxl
import boto3
from botocore.exceptions import ClientError
from botocore.config import Config
import sys
import datetime

    
def createDDBItem(tableName,json):
    table = dynamodb.Table(tableName)
    table.put_item(Item=json)

def updateDDBItem(tableName,keyJson,dict):
    print(f"Table Name: {tableName} Key: {keyJson} Update JSON: {dict}")
    table = dynamodb.Table(tableName)
    try:
        response = table.update_item(Key=keyJson,
                        UpdateExpression=dict[0], 
                        ExpressionAttributeValues=dict[1],
                        ReturnValues="UPDATED_NEW")
    except ClientError as err:
        print(err)
        print(f"Key {keyJson} not present in table {tableName}")
    else:
        return response["Attributes"]



def queryDDBItem(tableName,tableId,resultField):
    table = dynamodb.Table(tableName)
    try:
        response = table.get_item(
            Key = {'id':tableId},
            ProjectionExpression  = resultField
        )
        return response['Item']
    except:
        print("The key is not present")
        return ""

def prepareRetailerJson(retailerDict,eaUpdatesDict,mkUpdatesDict,brand):
    retailer={}
    eaJson={}
    mkJson={}
    if len(eaUpdatesDict)>0: 
        for key,val in eaUpdatesDict.items():
            if(key!='DEFAULT'):
                redAntFlag = val['RedAnt'] if val.get('RedAnt')!=None else "False"
                sfccFlag = val['SFCC'] if val.get('SFCC')!=None else "False"
                sfscFlag = val['SFSC'] if val.get('SFSC')!=None else "False"
                gbqFlag = val['GBQ'] if val.get('GBQ')!=None else "False"
                
                eaJson.update({
                    key : {
                            "sendOrderStatusToRedAnt":redAntFlag,                            
                            "sendOrderStatusToSFCC": sfccFlag,                            
                            "sendOrderStatusToSFSC": sfscFlag,                            
                            "sendOrderStatusToGBQ": gbqFlag,
                            "sendReturnOrderStatusToRedAnt": redAntFlag,
                            "sendReturnOrderStatusToSFCC": sfccFlag,
                            "sendReturnOrderStatusToSFSC": sfscFlag,
                            "sendReturnOrderStatusToGBQ": gbqFlag
                        }                            
                    }                    
                )
        eaJson.update({"sendOrderHistoryToXstore":  eaUpdatesDict['DEFAULT']['Xstore']})
    elif len(mkUpdatesDict)>0:
        for key,val in mkUpdatesDict.items():
            if(key!='DEFAULT'):
                redAntFlag = val['RedAnt'] if val.get('RedAnt')!=None else "False"
                sfccFlag = val['SFCC'] if val.get('SFCC')!=None else "False"
                sfscFlag = val['SFSC'] if val.get('SFSC')!=None else "False"
                gbqFlag = val['GBQ'] if val.get('GBQ')!=None else "False"
                mkJson.update({
                    key : {
                            "sendOrderStatusToRedAnt": redAntFlag,                            
                            "sendOrderStatusToSFCC":  sfccFlag,                            
                            "sendOrderStatusToSFSC":  sfscFlag,                            
                            "sendOrderStatusToGBQ":  gbqFlag,
                            "sendReturnOrderStatusToRedAnt":  redAntFlag,
                            "sendReturnOrderStatusToSFCC":  sfccFlag,
                            "sendReturnOrderStatusToSFSC":  sfscFlag,
                            "sendReturnOrderStatusToGBQ":  gbqFlag 
                    }                    
                })
        mkJson.update({"sendOrderHistoryToXstore":  mkUpdatesDict['DEFAULT']['Xstore']})
       
    retailer = {
        "id": "retailer_"+retailerDict['retailerId'],
        "brandName": {"sfsc_case": retailerDict['sfscCaseBrandName']},
        "retailer_id": retailerDict['retailerId'],
        "retailer_password": retailerDict['retailerPwd'],
        "retailer_username": retailerDict['retailerUsername'],
        "brand":brand
    }

    print(eaJson)
    print(mkJson)

    if len(eaUpdatesDict)>0:
        retailer.update({"sourceBasedRouting":{"endless_aisle":eaJson}})
    elif len(mkUpdatesDict)>0:
        if retailer["sourceBasedRouting"]!=None:
            retailer["sourceBasedRouting"].append({"marketplace":mkJson})
        else:
            retailer.update({"sourceBasedRouting":{"marketplace":mkJson}})

    return retailer

def updateEcomStoreEntries(storeDict,retailerDict,tableName):
    for key,value in storeDict.items():
        if(key == 'ecomStoreIds'):
            for item in value:         
                #print(prepareStoreJson(retailerDict,item["EcomCountryCode"],storeDict['brandName'],'N'))     
                updateDDBItem(tableName,{"id":item["id"]},prepareStoreJson(retailerDict,item["EcomCountryCode"],storeDict['brandName'],'N'))
        elif(key == 'storeIds'):
            for item in value:
                #print(prepareStoreJson(retailerDict,item["Country"],storeDict['brandName'],item["RedAnt"]))
                updateDDBItem(tableName,{"id":item["id"]},prepareStoreJson(retailerDict,item["Country"],storeDict['brandName'],item["RedAnt"]))

def prepareStoreJson(retailerDict,countryCode,brandName,redAnt):

    updateStoreExpr = 'SET inputData.retailer_id=:val1,inputData.retailer_password=:val2,inputData.retailer_username=:val3,' \
                        'inputData.fluent_OMS=:val4,inputData.hybrid_OMS=:val5,inputData.store=:val6,inputData.country_code=:val7'
    
    #expressionAttrValues= {'#s':'store'}
    
    expressionValues = {':val1':retailerDict['retailerId'],':val2':retailerDict['retailerPwd'],':val3':retailerDict['retailerUsername'],
                        ":val4":"Y",":val5":"N",":val6":brandName,":val7":countryCode}
    
    if(redAnt == 'Y'):
        updateStoreExpr += 'inputData.brandName=:val8,inputData.EA_enabled=:val9'
    
        expressionValues.update({':val8':brand,':val9':'Y'})
        
    return updateStoreExpr, expressionValues#,expressionAttrValues

def updateWmsStoreEntries(wmsDict,retailerDict,tableName):
    print(retailerDict)
    for key,value in wmsDict.items():
        for item in value:         
            print(prepareWmsStoreJson(retailerDict,item["wm9_facility"],item["wm9_storer"]))     
            #updateDDBItem(tableName,{"id":item["id"]},prepareWmsStoreJson(retailerDict,item["WmsCountryCode"],storeDict['brandName'],'N'))


def prepareWmsStoreJson(retailerDict,wm9Facility,wm9Storer):

    updateStoreExpr = 'SET inputData.retailer_id=:val1,inputData.retailer_password=:val2,inputData.retailer_username=:val3,' \
                        'inputData.Fluent_OMS=:val4,inputData.hybrid_OMS=:val5'
    
    
    expressionValues = {':val1':retailerDict['id'],':val2':retailerDict['retailer_password'],':val3':retailerDict['retailer_username'],
                        ":val4":"Y",":val5":"N"}
    
    if(wm9Facility!=None):
        updateStoreExpr += 'inputData.wm9_facility=:val6,inputData.wm9_storer=:val7'
        expressionValues.update({"val6":wm9Facility,"val7":wm9Storer})
    
    return updateStoreExpr, expressionValues

def updatePaymentEntries(paymentDict,tableName):
    for key in paymentDict.keys():
        #print(preparePaymentJson(paymentDict,key))
        createDDBItem(tableName,preparePaymentJson(paymentDict,key))

def preparePaymentJson(paymentDict,key):
    
    paymentCountryInfo = paymentDict[key]
    pspJson={}
    for item in paymentCountryInfo:
        pspKey=(brand+"."+item['Country']).lower()
        keyJson={}
        for paymentKeys in item['keys']: 
            for paymentKey,paymentKeyVal in paymentKeys.items():
                keyJson.update({paymentKey:paymentKeyVal})
        
        pspJson[pspKey]=keyJson

    paymentJson={
        "id":key,
        "inputData":pspJson
    }
    return paymentJson

def updateProductEntries(productDict,tableName1,tableName2):
    print("Update Product ENtries")
    productIdList = queryDDBItem(tableName1,"FL_P_D",'brands')
    if productDict['id'] in productIdList:
        print("Brand code is added in the FL_P_D entry")
    else:
        updateProductExpr = 'SET brands=:val1'        
        expressionValues = {':val1':productIdList['brands']+","+productDict['id']}
        updateDict = updateProductExpr, expressionValues
        updateDDBItem(tableName1,{"id":"FL_P_D"},updateDict)
        
    createDDBItem(tableName2,productDict)

def updateInventoryEntries(invDict,tableName):
    print("Update Inventory ENtries") 
    for key in invDict:
        createDDBItem(tableName,invDict[key])

def updateSiocsInventoryEntries(invDict,tableName1,tableName2):
    print("Update SIOCS  Entries") 
    for storeKey,storeVal in invDict.items():
        brand = storeVal[0]["brandName"]
        brandNames = queryDDBItem(tableName1,storeKey,'brandNames')
        if brand in brandNames:
            print(f"Brand code is added in the "+{}+" entry")
        else:
            if(not brandNames):
                brandJson={
                    "location":brand,
                    "brandNames":brand
                }
                createDDBItem(tableName1,brandJson)
            else:
                updateStoreExpr = 'SET brandNames=:val1'        
                expressionValues = {':val1':brandNames['brandNames']+","+brand}
                updateDict = updateStoreExpr, expressionValues
                updateDDBItem(tableName1,{"id":storeKey},updateDict)

        siocsJson={
            "brandName":brand,
            "fileProcessingDay":storeVal[1]["fileProcessingDay"],
            "location":storeKey,
            "outputType":"json"
        }   
        createDDBItem(tableName2,siocsJson)

def createCEConfig(ceConfig,geoConfig,tableName1,tableName2):
    for key in ceConfig:
        print(ceConfig[key])        
        createDDBItem(tableName1,ceConfig[key])
    for key in geoConfig:
        print(geoConfig[key])
        createDDBItem(tableName2,geoConfig[key])

def createCEOrderConfig(ceOrderConfig,tableName1,tableName2):
    orderCEKey = "FL_CE_Order"
    for key in ceOrderConfig:
        brandList = queryDDBItem(tableName1,orderCEKey,"brands")
        if key in brandList['brands']:
            print("Brand code is added in the FL_CE_Order entry")
        else:
            updateBrandExpr = 'SET brands=:val1'        
            expressionValues = {':val1':brandList['brands']+","+key}
            updateDict = updateBrandExpr, expressionValues
            updateDDBItem(tableName1,{"id":orderCEKey},updateDict)
        
        createDDBItem(tableName2,ceOrderConfig[key])
def createCEReturnOrderConfig(ceReturnOrderConfig,tableName1,tableName2):
    returnOrderCEKey = "FL_CE_Return"
    for key in ceReturnOrderConfig:
        brandList = queryDDBItem(tableName1,returnOrderCEKey,"brands")
        if key in brandList['brands']:
            print("Brand code is added in the FL_CE_Return entry")
        else:
            updateBrandExpr = 'SET brands=:val1'        
            expressionValues = {':val1':brandList['brands']+","+key}
            updateDict = updateBrandExpr, expressionValues
            updateDDBItem(tableName1,{"id":returnOrderCEKey},updateDict)
        
        createDDBItem(tableName2,ceReturnOrderConfig[key])
def createCEProductFeedConfig(ceProductFeedConfig,tableName1,tableName2):
    productKey="CE_P_D"
    for key in ceProductFeedConfig:
        brandList = queryDDBItem(tableName1,productKey,"brands")
        if key in brandList['brands']:
            print("Brand code is added in the CE_P_D entry")
        else:
            updateBrandExpr = 'SET brands=:val1'        
            expressionValues = {':val1':brandList['brands']+","+key}
            updateDict = updateBrandExpr, expressionValues
            updateDDBItem(tableName1,{"id":productKey},updateDict)
        
        createDDBItem(tableName2,ceProductFeedConfig[key])
def createCEPriceFeedConfig(cePriceFeedConfig,tableName):
    for key in cePriceFeedConfig:
        createDDBItem(tableName,cePriceFeedConfig[key])
def createCEInventoryFeedConfig(ceInventoryFeedConfig,tableName):
    for key in ceInventoryFeedConfig:
        createDDBItem(tableName,ceInventoryFeedConfig[key])

def main():
    # Read data from each sheet and iterate through the entries to create DDB entries

    wb = openpyxl.load_workbook("cosmos_setup.xlsx")
    retailerDict={}
    storeDict={}
    wmsDict={}
    eaUpdateDict={}
    mkUpdateDict={}
    paymentDict={}
    productDict={}
    invSfccDict={}
    invFarfetchDict={}
    invRedAntDict={}
    invSiocsDict={}
    ceConfig={}
    geoConfig={}
    ceOrderConfig={}
    ceReturnOrderConfig={}
    ceProductFeedConfig={}
    cePriceFeedConfig={}
    ceInventoryFeedConfig={}

    for sheet in wb:
        if(sheet.title=='Retailer'):
            storeDict.setdefault('storeIds', [])
            storeDict.setdefault('ecomStoreIds', [])
            wmsDict.setdefault('wmsStore',[])
            for row in sheet.iter_rows(min_row=2,max_row=50,values_only=True):
                if row[0] != None:
                    retailerDict['retailerId'] = str(row[0])
                if row[1] != None:
                    retailerDict['retailerUsername'] = row[1]
                if row[2] != None:
                    retailerDict['retailerPwd'] = row[2]
                if row[3] != None:
                    retailerDict['sfscCaseBrandName'] = row[3]
                if row[4] != None and row[6] != None and row[5]!=None:
                    storeDict['storeIds'].append({"id":str(row[4]),"RedAnt":row[6],"Country":row[5]})
                if row[9] != None:
                    storeDict['brandName'] = row[9]
                    brand=row[9]
                if row[7] != None:
                    storeDict['ecomStoreIds'].append({"id":str(row[7]),"EcomCountryCode":row[8]})
                if row[10] != None:
                    wmsDict['wmsStore'].append({"id":str(row[10]),"WmsCountryCode":row[11],"wm9_facility":row[12],"wm9_storer":row[13]})
        if(sheet.title=='EA_StatusUpdates'):        
            for row in sheet.iter_rows(min_row=2,max_row=10,values_only=True):
                if(row[0]!=None and row[0]!='DEFAULT'):
                    val1 = sheet.cell(row=1,column=2).value
                    val2 = sheet.cell(row=1,column=3).value
                    val3 = sheet.cell(row=1,column=4).value
                    val4 = sheet.cell(row=1,column=5).value
                    eaUpdateDict[row[0]] = {val1:row[1],val2:row[2],val3:row[3],val4:row[4]}
                if(row[0] =='DEFAULT'):                
                    val5 = sheet.cell(row=1,column=6).value
                    eaUpdateDict[row[0]] = {val5:row[5]}
        if(sheet.title=='MK_StatusUpdates'):        
            for row in sheet.iter_rows(min_row=2,max_row=10,values_only=True):
                if(row[0]!=None and row[0]!='DEFAULT'):
                    val1 = sheet.cell(row=1,column=2).value
                    val2 = sheet.cell(row=1,column=3).value
                    val3 = sheet.cell(row=1,column=4).value
                    val4 = sheet.cell(row=1,column=5).value
                    mkUpdateDict[row[0]]= {val1:row[1],val2:row[2],val3:row[3],val4:row[4]}
                if(row[0] =='DEFAULT'):                
                    val5 = sheet.cell(row=1,column=6).value
                    mkUpdateDict[row[0]]= {val5:row[5]}
        if(sheet.title=='Checkout'): 
            paymentDict.setdefault(('checkout_'+brand).upper(),[])
            for rowVal in sheet.iter_rows(min_row=2,max_row=20,values_only=True):
                keyVals=[]
                if(rowVal[1]!=None and rowVal[2]!=None and rowVal[3]!=None and rowVal[4]!=None):
                    keyVals.append({sheet.cell(row=1,column=2).value:rowVal[1]})
                    keyVals.append({sheet.cell(row=1,column=3).value:rowVal[2]})
                    keyVals.append({sheet.cell(row=1,column=4).value:rowVal[3]})
                    keyVals.append({sheet.cell(row=1,column=5).value:rowVal[4]})
                    countryJson = {'Country':rowVal[0],'keys':keyVals}
                    paymentDict[('checkout_'+brand).upper()].append(countryJson)
        if(sheet.title=='Tabby'): 
            paymentDict.setdefault(('tabby_'+brand).upper(),[])
            for rowVal in sheet.iter_rows(min_row=2,max_row=20,values_only=True):
                keyVals=[]
                if(rowVal[1]!=None):
                    keyVals.append({sheet.cell(row=1,column=2).value:rowVal[1]})
                    countryJson = {'Country':rowVal[0],'keys':keyVals}
                    paymentDict[('tabby_'+brand).upper()].append(countryJson)
        if(sheet.title=='Tamara'): 
            paymentDict.setdefault(('tamara_'+brand).upper(),[])
            for rowVal in sheet.iter_rows(min_row=2,max_row=20,values_only=True):
                keyVals=[]
                if(rowVal[1]!=None and rowVal[2]!=None):
                    keyVals.append({sheet.cell(row=1,column=2).value:rowVal[1]})
                    keyVals.append({sheet.cell(row=1,column=3).value:rowVal[2]})
                    countryJson = {'Country':rowVal[0],'keys':keyVals}
                    paymentDict[('tamara_'+brand).upper()].append(countryJson)
        if(sheet.title=='Product'): 
            destinationDetails=[]
            sourceDetails=[]
            for column in sheet.iter_cols():
                column_name = column[0].value
                for i,cell in enumerate(column):
                    if(i==0):
                        continue
                    if cell!=None and type(cell.value)==datetime.datetime:
                        productDict[column_name]=str(cell.value)                        
                    elif cell!=None and column_name.startswith('destinationDetails'):
                        destinationDetails.append({column_name[len("destinationDetails_"):]:cell.value})                        
                    elif cell!=None and column_name.startswith('sourceDetails'):
                        sourceDetails.append({column_name[len("sourceDetails_"):]:cell.value})
                    else:
                        productDict[column_name]=cell.value
            
            destJson = {}
            for dest in destinationDetails:
                destJson.update(dest)
            sourceJson = {}
            for source in sourceDetails:
                sourceJson.update(source)
            productDict["destinationDetails"] = destJson
            productDict["sourceDetails"] = sourceJson

        if(sheet.title=='Inventory-SFCC'): 
            countryJson={}
            destinationDetails=[]
            sourceDetails=[]
            countryKey=""
            for row in sheet.iter_rows(min_row=2):
                for i,column in enumerate(sheet.iter_cols()):
                    column_name = column[0].value
                    if(column_name=='Country'):
                        countryKey = row[i].value
                        invSfccDict.setdefault(countryKey,{})
                    elif cell!=None and column_name.startswith('destinationDetails'):
                        destinationDetails.append({column_name[len("destinationDetails_"):]:row[i].value})
                    elif cell!=None and column_name.startswith('sourceDetails'):
                        sourceDetails.append({column_name[len("sourceDetails_"):]:row[i].value})
                    else:
                        invSfccDict.get(countryKey).update({column_name:row[i].value})
                destJson = {}
                for dest in destinationDetails:
                    destJson.update(dest)
                sourceJson = {}
                for source in sourceDetails:
                    sourceJson.update(source)
                invSfccDict.get(countryKey).update({"destinationDetails":destJson})
                invSfccDict.get(countryKey).update({"sourceDetails":sourceJson})           
 
        if(sheet.title=='Inventory-Farfetch'): 
            countryJson={}
            destinationDetails=[]
            sourceDetails=[]
            countryKey=""
            for row in sheet.iter_rows(min_row=2):
                for i,column in enumerate(sheet.iter_cols()):
                    column_name = column[0].value
                    if(column_name=='Country'):
                        countryKey = row[i].value
                        invFarfetchDict.setdefault(countryKey,{})
                    elif cell!=None and column_name.startswith('destinationDetails'):
                        destinationDetails.append({column_name[len("destinationDetails_"):]:row[i].value})
                    elif cell!=None and column_name.startswith('sourceDetails'):
                        sourceDetails.append({column_name[len("sourceDetails_"):]:row[i].value})
                    else:
                        invFarfetchDict.get(countryKey).update({column_name:row[i].value})
                destJson = {}
                for dest in destinationDetails:
                    destJson.update(dest)
                sourceJson = {}
                for source in sourceDetails:
                    sourceJson.update(source)
                invFarfetchDict.get(countryKey).update({"destinationDetails":destJson})
                invFarfetchDict.get(countryKey).update({"sourceDetails":sourceJson})       
        if(sheet.title=='Inventory-RedAnt'): 
            countryJson={}
            destinationDetails=[]
            sourceDetails=[]
            countryKey=""
            for row in sheet.iter_rows(min_row=2):
                for i,column in enumerate(sheet.iter_cols()):
                    column_name = column[0].value
                    if(column_name=='Country'):
                        countryKey = row[i].value
                        invRedAntDict.setdefault(countryKey,{})
                    elif cell!=None and column_name.startswith('destinationDetails'):
                        destinationDetails.append({column_name[len("destinationDetails_"):]:row[i].value})
                    elif cell!=None and column_name.startswith('sourceDetails'):
                        sourceDetails.append({column_name[len("sourceDetails_"):]:row[i].value})
                    else:
                        invRedAntDict.get(countryKey).update({column_name:row[i].value})
                destJson = {}
                for dest in destinationDetails:
                    destJson.update(dest)
                sourceJson = {}
                for source in sourceDetails:
                    sourceJson.update(source)
                invRedAntDict.get(countryKey).update({"destinationDetails":destJson})
                invRedAntDict.get(countryKey).update({"sourceDetails":sourceJson})    

        if(sheet.title=='Inventory-SIOCS'): 
            for row in sheet.iter_rows(min_row=2):
                invSiocsDict[row[0].value]=[]
                for i,col in enumerate(sheet.iter_cols(min_col=2)):
                    colName = col[0].value
                    invSiocsDict[row[0].value].append({colName:row[i+1].value})
        
        if(sheet.title=='CE'): 
            for row in sheet.iter_rows(min_row=2):
                ceConfig[row[0].value]={}
                for i,col in enumerate(sheet.iter_cols(min_col=1)):
                    colName = col[0].value
                    ceConfig[row[0].value].update({colName:row[i].value})

        if(sheet.title=='GeoCode'): 
            for row in sheet.iter_rows(min_row=2):
                geoConfig[row[0].value]={}
                for i,col in enumerate(sheet.iter_cols(min_col=1)):
                    colName = col[0].value
                    geoConfig[row[0].value].update({colName:row[i].value})

        if(sheet.title=='CEOrder'): 
            for row in sheet.iter_rows(min_row=2):
                ceOrderConfig[row[0].value]={}
                marketPlaceCols = []
                for i,col in enumerate(sheet.iter_cols(min_col=1)):
                    colName = col[0].value
                    if "_" in colName:
                        nameArr = colName.split("_")
                        marketplace = nameArr[0]
                        destnColumn = nameArr[1]
                        marketPlaceCols.append({destnColumn:row[i].value})
                    else:
                        ceOrderConfig[row[0].value].update({colName:row[i].value})
                sourceJson = {}
                for source in marketPlaceCols:
                    sourceJson.update(source)
                ceOrderConfig.get(row[0].value).update({marketplace:sourceJson})

        if(sheet.title=='CEReturn'): 
            for row in sheet.iter_rows(min_row=2):
                ceReturnOrderConfig[row[0].value]={}
                marketPlaceCols = []
                for i,col in enumerate(sheet.iter_cols(min_col=1)):
                    colName = col[0].value
                    if "_" in colName:
                        nameArr = colName.split("_")
                        marketplace = nameArr[0]
                        destnColumn = nameArr[1]
                        marketPlaceCols.append({destnColumn:row[i].value})
                    else:
                        ceReturnOrderConfig[row[0].value].update({colName:row[i].value})
                sourceJson = {}
                for source in marketPlaceCols:
                    sourceJson.update(source)
                ceReturnOrderConfig[row[0].value].update({marketplace:sourceJson})

        if(sheet.title=='CEProduct'): 
            for row in sheet.iter_rows(min_row=2):
                ceProductFeedConfig[row[0].value]={}
                childCols=[]
                for i,col in enumerate(sheet.iter_cols(min_col=1)):
                    colName = col[0].value
                    if "_" in colName: 
                        nameArr = colName.split("_")
                        header = nameArr[0]
                        childCol = nameArr[1]
                        childCols.append({childCol:row[i].value})
                    elif row[i].value!=None and type(row[i].value)==datetime.datetime:
                        ceProductFeedConfig[row[0].value].update({colName:str(row[i].value)})
                    elif row[i].value!=None:
                        ceProductFeedConfig[row[0].value].update({colName:row[i].value})
                sourceJson = {}
                for source in childCols:
                    sourceJson.update(source)
                ceProductFeedConfig[row[0].value].update({header:sourceJson})
        
        if(sheet.title=='CEPrice'): 
            for row in sheet.iter_rows(min_row=2):
                cePriceFeedConfig[row[0].value]={}
                childCols=[]
                for i,col in enumerate(sheet.iter_cols(min_col=1)):
                    colName = col[0].value
                    if "_" in colName: 
                        nameArr = colName.split("_")
                        header = nameArr[0]
                        childCol = nameArr[1]
                        childCols.append({childCol:row[i].value})
                    elif row[i].value!=None and type(row[i].value)==datetime.datetime:
                        cePriceFeedConfig[row[0].value].update({colName:str(row[i].value)})
                    elif row[i].value!=None:
                        cePriceFeedConfig[row[0].value].update({colName:row[i].value})
                sourceJson = {}
                for source in childCols:
                    sourceJson.update(source)
                cePriceFeedConfig[row[0].value].update({header:sourceJson})

        if(sheet.title=='CEInventory'): 
            for row in sheet.iter_rows(min_row=2):
                ceInventoryFeedConfig[row[0].value]={}
                childCols=[]
                for i,col in enumerate(sheet.iter_cols(min_col=1)):
                    colName = col[0].value
                    if "_" in colName: 
                        nameArr = colName.split("_")
                        header = nameArr[0]
                        childCol = nameArr[1]
                        childCols.append({childCol:row[i].value})
                    elif row[i].value!=None and type(row[i].value)==datetime.datetime:
                        ceInventoryFeedConfig[row[0].value].update({colName:str(row[i].value)})
                    elif row[i].value!=None:
                        ceInventoryFeedConfig[row[0].value].update({colName:row[i].value})
                sourceJson = {}
                for source in childCols:
                    sourceJson.update(source)
                ceInventoryFeedConfig[row[0].value].update({header:sourceJson})

    commandArgs = sys.argv[1:]
    for args in commandArgs:
        if args in ("dev","test","uat"):
            env = args
        if args in ("r","retailer"):
            print("Create or update retailer details")
            tableName = 'fluent-config-'+env
            #print(prepareRetailerJson(retailerDict,eaUpdateDict,mkUpdateDict,brand))
            createDDBItem(tableName,prepareRetailerJson(retailerDict,eaUpdateDict,mkUpdateDict))
            fluentDict = {brand:retailerDict['retailerId']}
            updateDDBItem(tableName,{"id":{"S":"fluent_brands"}},fluentDict)    
        elif args in ("s","store"):
            print("Update store details")
            tableName = 'store-lookup-'+env
            updateEcomStoreEntries(storeDict,retailerDict,tableName)
        elif args in ("psp","payment"):
            print("Create or Update payment details")
            tableName = 'payment-master-'+env
            updatePaymentEntries(paymentDict,tableName)
        elif args in ("p","product"):
            print("Create or Update product details")
            tableName1 = 'product-batch-process-'+env
            tableName2 = 'product-brand-details-'+env
            updateProductEntries(productDict,tableName1,tableName2)
        elif args in ("sfcc-inv","SFCC inventory"):
            print("Create or Update SFCC inventory details")
            tableName = 'product-brand-details-'+env
            updateInventoryEntries(invSfccDict,tableName)
        elif args in ("farfetch-inv","Farfetch inventory"):
            print("Create or Update Farfetch inventory details")
            tableName = 'product-brand-details-'+env
            updateInventoryEntries(invFarfetchDict,tableName)
        elif args in ("redant-inv","RedAnt inventory"):
            print("Create or Update RedAnt inventory details")
            tableName = 'product-brand-details-'+env
            updateInventoryEntries(invRedAntDict,tableName)
        elif args in ("siocs-inv","SIOCS inventory"):
            print("Create or Update SIOCS inventory details")
            tableName1 = 'siocs-location-brand-snapshot-'+env
            tableName2 = 'siocs-brand-dtl-snapshot-'+env
            updateSiocsInventoryEntries(invSiocsDict,tableName1,tableName2)
        elif args in ("wms","warehouse"):
            print("Create or Update warehouse details")
            tableName = 'wh-fulfilment-'+env
            updateWmsStoreEntries(wmsDict,prepareRetailerJson(retailerDict,eaUpdateDict,mkUpdateDict),tableName)
        elif args in ("ceConfig"):
            print("Create or Update ce config details")
            tableName1 = 'ce-config-'+env
            tableName2 = 'geocoding-config-'+env
            createCEConfig(ceConfig,geoConfig,tableName1,tableName2)
        elif args in ("ceOrder"):
            print("Create or Update ce order config details")
            tableName1 = 'product-batch-process-'+env
            tableName2 = 'product-brand-details-'+env
            createCEOrderConfig(ceOrderConfig,tableName1,tableName2)
        elif args in ("ceReturnOrder"):
            print("Create or Update ce return order config details")
            tableName1 = 'product-batch-process-'+env
            tableName2 = 'product-brand-details-'+env
            createCEReturnOrderConfig(ceReturnOrderConfig,tableName1,tableName2)
        elif args in ("ceProduct"):
            print("Create or Update ce product feed details")
            tableName1 = 'product-batch-process-'+env
            tableName2 = 'product-brand-details-'+env
            createCEProductFeedConfig(ceProductFeedConfig,tableName1,tableName2)
        elif args in ("cePrice"):
            print("Create or Update ce price feed details")
            tableName = 'product-brand-details-'+env
            createCEPriceFeedConfig(cePriceFeedConfig,tableName)
        elif args in ("ceInventory"):
            print("Create or Update ce inventory feed details")
            tableName = 'product-brand-details-'+env
            createCEInventoryFeedConfig(ceInventoryFeedConfig,tableName)

    print(f"Environment",env)


if __name__ == "__main__":
    
    #Setup DynamoDB connection
    
    config = Config(
        retries={'max_attempts': 3, 'mode': 'standard'}
    )
    dynamodb = boto3.resource('dynamodb',config=config)
    brand = ''
    env="dev"
    main()